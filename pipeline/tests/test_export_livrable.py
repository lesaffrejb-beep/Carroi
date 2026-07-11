"""Tests du livrable client et de l'archive de l'actif (décisions op docs/16).

(a) 40_export_client.ecrire_xlsx : le .xlsx de travail (en-tête figé, filtres, largeurs).
(b) common.archiver_copie_datee : copie datée de data/final/ (point-zéro des millésimes).
"""
from __future__ import annotations

import importlib
import logging
from datetime import date

import pandas as pd
import pytest
from openpyxl import load_workbook

import common

export = importlib.import_module("40_export_client")
log = logging.getLogger("tests")


def _livrable(n=5):
    return pd.DataFrame({
        "id_ban": [f"ban_{i}" for i in range(n)],
        "adresse": [f"{i} rue de la Piscine" for i in range(n)],
        "code_postal": ["49000"] * n,
        "commune": ["Angers"] * n,
        "lat": [47.47 + i / 1000 for i in range(n)],
        "lon": [-0.55 - i / 1000 for i in range(n)],
        "surface_m2": [30.0 + i for i in range(n)],
        "confiance": ["haute"] * n,
    })


# ------------------------------------------------------------- (a) xlsx

def test_xlsx_contenu_et_mise_en_forme(tmp_path):
    df = _livrable(5)
    out = tmp_path / "piscines_zone.xlsx"
    export.ecrire_xlsx(df, out, feuille="piscines")

    assert out.exists()
    wb = load_workbook(out)
    assert "piscines" in wb.sheetnames
    ws = wb["piscines"]

    # En-tête = colonnes du df ; une ligne de données par adresse.
    entete = [c.value for c in ws[1]]
    assert entete == list(df.columns)
    assert ws.max_row == len(df) + 1, "une ligne d'en-tête + une par adresse"

    # En-tête figé + filtres actifs (docs/16 §1).
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == ws.dimensions

    # Largeurs ajustées : la colonne 'adresse' (valeurs longues) plus large que 'confiance'.
    from openpyxl.utils import get_column_letter
    larg = {col: ws.column_dimensions[get_column_letter(i)].width
            for i, col in enumerate(df.columns, start=1)}
    assert larg["adresse"] > larg["confiance"] > 0


def test_xlsx_ne_deforme_pas_les_valeurs(tmp_path):
    df = _livrable(3)
    out = tmp_path / "x.xlsx"
    export.ecrire_xlsx(df, out)
    ws = load_workbook(out).active
    # Première ligne de données correspond bien au df.
    ligne1 = [c.value for c in ws[2]]
    assert ligne1[0] == "ban_0"
    assert ligne1[1] == "0 rue de la Piscine"
    assert ligne1[ list(df.columns).index("surface_m2") ] == 30.0


# ------------------------------------------------------ (b) archive datée

def _cfg_final(tmp_path):
    return {"paths": {"final": str(tmp_path / "final")}}


def _parquet_bidon(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame({"id_ban": ["ban_1"], "surface_m2": [30.0]}).to_parquet(path)
    return path


def test_archive_cree_une_copie_datee(tmp_path):
    cfg = _cfg_final(tmp_path)
    src = _parquet_bidon(tmp_path / "final" / "piscines_qualifiees_49.parquet")

    dest = common.archiver_copie_datee(src, "piscines", "49", cfg, log)

    attendu = f"piscines_49_{date.today().isoformat()}.parquet"
    assert dest.name == attendu
    assert dest.parent.name == "archive"
    assert dest.exists()
    assert dest.read_bytes() == src.read_bytes(), "copie fidèle de l'actif"


def test_archive_idempotente_le_meme_jour(tmp_path):
    cfg = _cfg_final(tmp_path)
    src = _parquet_bidon(tmp_path / "final" / "piscines_qualifiees_49.parquet")
    d1 = common.archiver_copie_datee(src, "piscines", "49", cfg, log)
    d2 = common.archiver_copie_datee(src, "piscines", "49", cfg, log)
    assert d1 == d2, "une re-génération le même jour réécrit la même copie (pas de doublon)"


def test_archive_echoue_bruyamment_si_source_absente(tmp_path):
    cfg = _cfg_final(tmp_path)
    with pytest.raises(FileNotFoundError):
        common.archiver_copie_datee(tmp_path / "inexistant.parquet", "piscines", "49", cfg, log)
